import os
import sys
import time
import Monitering.Query
from PyQt5.QtWidgets import QFileDialog, QDialog
from PyQt5.QtCore import *
from openpyxl import load_workbook

print(sys.version, "\n")
try:

    from watchdog.observers import Observer

    from watchdog.events import FileSystemEventHandler

except ModuleNotFoundError as e:

    print (e)

    os.system("pip install watchdog")


from PyQt5 import QtCore, QtWidgets





class Ui_MainWindow(object):

    path1 = ""
    path2 = ""
    temp_src = ""



    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(419, 192)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.FolderButton1 = QtWidgets.QPushButton(self.centralwidget)
        self.FolderButton1.setGeometry(QtCore.QRect(320, 10, 71, 23))
        self.FolderButton1.setObjectName("FolderButton1")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(10, 12, 101, 16))
        self.label.setObjectName("label")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(110, 10, 201, 21))
        self.lineEdit.setObjectName("lineEdit")
        self.RunButton_6 = QtWidgets.QPushButton(self.centralwidget)
        self.RunButton_6.setGeometry(QtCore.QRect(150, 90, 111, 41))
        self.RunButton_6.setObjectName("RunButton_6")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(10, 50, 101, 31))
        self.label_6.setObjectName("label_6")
        self.lineEdit_6 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_6.setGeometry(QtCore.QRect(110, 60, 201, 21))
        self.lineEdit_6.setObjectName("lineEdit_6")
        self.FolderButton5_2 = QtWidgets.QPushButton(self.centralwidget)
        self.FolderButton5_2.setGeometry(QtCore.QRect(320, 60, 71, 23))
        self.FolderButton5_2.setObjectName("FolderButton5_2")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 419, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.dialog = QDialog(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.FolderButton1.setText(_translate("MainWindow", "폴더열기"))
        self.label.setText(_translate("MainWindow", "모니터링 폴더"))
        self.RunButton_6.setText(_translate("MainWindow", "Run"))
        self.label_6.setText(_translate("MainWindow", "생성할 폴더"))
        self.FolderButton5_2.setText(_translate("MainWindow", "TXT 비교"))

        #기본세팅 및 버튼 연결
        path1 = Monitering.Query.show_part(self)
        print(path1[0][0])
        self.lineEdit.setText(path1[0][0])
        Ui_MainWindow.path1 = path1[0][0]
        self.FolderButton1.clicked.connect(self.Folder_Open)
        self.RunButton_6.clicked.connect(self.Run_PCR)
        self.FolderButton5_2.clicked.connect(self.Txt)
        self.lineEdit.setDisabled(True)

    def Folder_Open(self):
        QtWidgets.QMessageBox.information(MainWindow, "경로선택", "경로를 선택해주세요")
        path1 = QFileDialog.getExistingDirectory()
        path2 = os.path.realpath(path1)
        path2 = path1
        Ui_MainWindow.path2 = path2
        self.lineEdit.setText(path2)

    # PCR Run 버튼
    def Run_PCR(self):
        if Ui_MainWindow.path2 == "":
            myWatcher = Watcher(Ui_MainWindow.path1)
            myWatcher.run()
        else:
            myWatcher = Watcher(Ui_MainWindow.path2)
            myWatcher.run()


    # def Excel(self):
    #     # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    #
    #     load_wb = load_workbook("C:\shinhoo\Monitering\과일.xlsx", data_only=True)
    #     # 시트 이름으로 불러오기
    #     load_ws = load_wb['Sheet1']
    #
    #     # 셀 주소로 값 출력
    #     print(load_ws['A1'].value)
    #
    #     # 셀 좌표로 값 출력
    #     print(load_ws.cell(1, 2).value)

    def Txt(self):
        #데이터베이스에 있는 텍스트파일
        temp = Monitering.Query.Sel_Bcd(self)
        file = open(temp[0][0], "r", encoding="utf8")
        #현재 비교할 텍스트 파일
        temp2 = Monitering.Query.Sel_Bcd2(self)
        file2 = open(temp2[0][0], "r", encoding="utf8")


        lines = file.readlines()# list 형태로 읽어옴
        lines2 = file2.readlines()
        cnt = 0




        for line in lines:
            if not line:
                break
            temp3 = lines[cnt].split()
            temp4 = lines2[cnt].split()

            if temp3[2] == temp4[2]:

                print("{0}번째 해당 줄이 같습니다.".format(cnt))
                print(temp3[2])

            elif temp3[2] != temp4[2]:
                print("{0}번째 해당 줄이 다릅니다.".format(2))
                print("1번 리스트 : " + temp3[2])
                print("2번 리스트 : " + temp4[2])
            cnt += 1

        file.close()
        file2.close()

class Handler(FileSystemEventHandler):

    def on_created(self, event): # 파일 생성시

        Ui_MainWindow.temp_src = "A"

        import sys
        app = QtWidgets.QApplication(sys.argv)
        MainWindow = QtWidgets.QMainWindow()
        ui = Ui_MainWindow()
        ui.setupUi(MainWindow)




        temp = event.src_path
        temp = temp.replace('\\','/')

        print (f'event type : {event.event_type}\n'
               f'event src_path : {event.src_path}')


        Monitering.Query.update(self, temp)


        if event.is_directory:
            print ("디렉토리 생성")
        else: # not event.is_directory

            """
            Fname : 파일 이름
            Extension : 파일 확장자 
            """
            Fname, Extension = os.path.splitext(os.path.basename(event.src_path))
            '''
             1. zip 파일

             2. exe 파일

             3. lnk 파일

            '''

            if Extension == '.zip':

                print (".zip 압축 파일 입니다.")

            elif Extension == '.exe':

                print (".exe 실행 파일 입니다.")

                os.remove(Fname + Extension)   # _파일 삭제 event 발생

            elif Extension == '.lnk':
                print (".lnk 링크 파일 입니다.")




    def on_deleted(self, event):

        print("삭제 이벤트 발생")
        print(f'event type : {event.event_type}\n'
              f'event src_path : {event.src_path}')

    def on_moved(self, event): # 파일 이동시

        print("업데이트 이벤트 발생")
        # print (f'event type : {event.event_type}\n')
        print(f'event type : {event.event_type}\n'
              f'event src_path : {event.src_path}')
class Watcher:
    # 생성자
    def __init__(self, path):

        print ("감시 중 ...")

        self.event_handler = None      # Handler

        self.observer = Observer()     # Observer 객체 생성

        self.target_directory = path   # 감시대상 경로

        self.currentDirectorySetting() # instance method 호출 func(1)

        Ui_MainWindow.temp_src == "AA"



    # func (1) 현재 작업 디렉토리

    def currentDirectorySetting(self):

        print ("====================================")

        print ("현재 작업 디렉토리:  ", end=" ")

        os.chdir(self.target_directory)

        print ("{cwd}".format(cwd = os.getcwd()))

        print ("====================================")



    # func (2)

    def run(self):

        self.event_handler = Handler() # 이벤트 핸들러 객체 생성

        self.observer.schedule(
            self.event_handler,
            self.target_directory,
            recursive=True

        )

        self.observer.start() # 감시 시작

        try:

            while True: # 무한 루프
                time.sleep(1) # 1초 마다 대상 디렉토리 감시
                # self.observer.stop()  # 감시 중단
                if Ui_MainWindow.temp_src == "A":
                    # print("감시 중지...")
                    # self.observer.stop()  # 감시 중단
                    # print("Error")
                    # self.observer.join()
                    break

        except KeyboardInterrupt: # 사용자에 의해 "ctrl + z" 발생시

            print ("감시 중지...")
            self.observer.stop() # 감시 중단
            print("Error")
            self.observer.join()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

    sys.exit(app.exec_())

# myWatcher2 = Watcher("C:/shinhoo\20201125180621.txt")
# myWatcher.run()

# myWatcher2.run()

